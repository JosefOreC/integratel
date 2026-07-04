"""
service_reports.py
Genera archivos CSV y Excel con los datos de clientes para exportación.
"""

import csv
import io
from app.services.service_clients import get_clients

HEADERS = ["ID Cliente", "Segmento", "Departamento", "Antigüedad (meses)", "ARPU (S/.)", "Estado", "Churn"]


def _rows() -> list[list]:
    clients = get_clients()
    return [
        [
            c["id_cliente"],
            c["segmento"],
            c["departamento"],
            c["antiguedad_meses"],
            round(c["arpu"], 2),
            c["estado"],
            "Sí" if c["churn"] == 1 else "No",
        ]
        for c in clients
    ]


def generate_csv() -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(HEADERS)
    writer.writerows(_rows())
    return output.getvalue()


def generate_excel() -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clientes"

    # Cabecera estilizada
    header_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(_rows(), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Ajustar ancho de columnas
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
